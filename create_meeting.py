import requests
import time
import openpyxl as xl
import datetime
import os, sys

# replace with your client ID
client_id = os.environ.get("CLIENT_ID") 

# replace with your account ID
account_id = os.environ.get("ACCOUNT_ID")

# replace with your client secret
client_secret = os.environ.get("CLIENT_SECRET") 

# print(f"client id is {client_id}")
# print(f"client secret  is {client_secret}")
# print(f"account id is {account_id}")

auth_token_url = "https://zoom.us/oauth/token"
api_base_url = "https://api.zoom.us/v2"


# Get access token function
def get_access_token():
    """Retrieves an access token, handling expiration and reuse."""

    access_token = None
    issued_time = None

    while True:
        if access_token and not has_token_expired(access_token, issued_at=issued_time):
            return access_token

        data = {
            "grant_type": "account_credentials",
            "account_id": account_id,
            "client_secret": client_secret
        }
        response = requests.post(auth_token_url, auth=(client_id, client_secret), data=data)

        if response.status_code != 200:
            raise Exception("Unable to get access token")

        response_data = response.json()
        access_token = response_data["access_token"]
        issued_time = time.time()
        return access_token

# Check if the access_token is expired
def has_token_expired(access_token, issued_at=None):
    """
    Checks if the access token has expired based on its issued time and validity period.

    Args:
    access_token: The access token string.
    issued_at: (Optional) The timestamp of the token issuance (Unix time).

    Returns:
    True if the token has expired, False otherwise.
    """

    # Validity period for Zoom access tokens is 1 hour (3600 seconds)
    token_validity = 3300

    # Use provided issued_at or current time if not given
    if not issued_at:
        issued_at = time.time()

    # Check if token age exceeds validity period with a safe margin
    # to account for potential time skews between systems
    return time.time() - issued_at > token_validity + 60


# create the Zoom link function
def create_meeting(meeting_name, time, account, duration, type):
        
    if int(type) == 2:
        payload = {
        "default_password": "false",
        "duration": duration,
        "password": "12345",
        "settings": {
            "allow_multiple_devices": "false",
            "approval_type": 0,
            "audio": "voip",
            "close_registration": "false",
            "contact_email": account,
            "contact_name": "Isha Foundation",
            "encryption_type": "enhanced_encryption",
            "focus_mode": "true",
            "host_video": "false",
            "mute_upon_entry": "true",
            "participant_video": "true",
            "registrants_confirmation_email": "false",
            "registrants_email_notification": "false",
            "show_share_button": "false",
            "waiting_room": "true",
            "continuous_meeting_chat": {
            "enable": "true",
            },
        },
        "start_time": time,
        "timezone": "Asia/Calcutta",
        "topic": meeting_name,
        "type": type
        }

    if int(type) == 3:
        payload = {
        "default_password": "false",
        "password": "12345",
        "settings": {
            "allow_multiple_devices": "false",
            "audio": "voip",
            "contact_email": account,
            "contact_name": "Isha Foundation",
            "encryption_type": "enhanced_encryption",
            "focus_mode": "true",
            "host_video": "false",
            "mute_upon_entry": "true",
            "participant_video": "true",
            "waiting_room": "true",
            "continuous_meeting_chat": {
            "enable": "true",
            },
        },
        "timezone": "Asia/Calcutta",
        "topic": meeting_name,
        "type": type
        }


    
    access_token = get_access_token()
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    resp = requests.post(f"{api_base_url}/users/{account}/meetings", 
                            headers=headers, 
                            json=payload)
    
    if resp.status_code!=201:
        print("Unable to generate meeting link")
    response_data = resp.json()
    
    return response_data
    


def create_meetings_from_excel():
    input_workbook = xl.load_workbook("meeting_names.xlsx")  # Adjust file path
    input_sheet = input_workbook["Sheet1"]  # Adjust sheet name

    output_workbook = xl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.cell(row=1, column=1).value = "Meeting Name"
    output_sheet.cell(row=1, column=2).value = "Meeting Link"
    output_sheet.cell(row=1, column=3).value = "Meeting ID"
    output_sheet.cell(row=1, column=4).value = "Passcode"

    row_num = 2  # Start writing from row 2
    for row in input_sheet.iter_rows(min_row=2):  # Skip header row
        meeting_name = row[1].value  # Assuming meeting names are in column A
        account = row[3].value # Email address in which meeting needs to be created
        if row[5].value == "registration":
            type = 2
            duration = int(row[4].value)
            meeting_date = row[2].value.isoformat()  # Existing datetime object
        elif row[5].value == "recurring":
            type = 3
        else:
            type = 3
        
        try:
            meeting_details = create_meeting(meeting_name, meeting_date, account, duration, type)
            meeting_name = meeting_details["topic"]
            meeting_link = meeting_details["join_url"]
            meeting_id = meeting_details["id"]
            meeting_passcode = meeting_details["password"]
            output_sheet.cell(row=row_num, column=1).value = meeting_name  # Write link to column B
            output_sheet.cell(row=row_num, column=2).value = meeting_link  # Write link to column B
            output_sheet.cell(row=row_num, column=3).value = meeting_id  # Write meeting ID  to column C
            output_sheet.cell(row=row_num, column=4).value = meeting_passcode  # Write link to column D
            print(f"Meeting {meeting_name} with meeting id {meeting_id} is created successfully")
            row_num += 1
        except Exception as e:
            print(f"Error creating meeting for {meeting_name}: {e}")

    output_workbook.save("meeting_links.xlsx")  # Adjust output file path

create_meetings_from_excel()


# delete a zoom meeting
def delete_meeting(meetingId):
    access_token = get_access_token()
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }
    resp = requests.delete(f"{api_base_url}/meetings/{meetingId}", headers=headers)
                            
    
    if resp.status_code == 204:
        print(f"Meeting with Id: {meetingId} deleted successfully")
    else:
        print("Failed to delete meeting:", resp.text)
    
        
def delete_meeting_from_excel():
     input_workbook = xl.load_workbook("meeting_links.xlsx")  # Adjust file path
     input_sheet = input_workbook["Sheet"]  # Adjust sheet name
     for row in input_sheet.iter_rows(min_row=2):
        meetingId = row[2].value
        delete_meeting(meetingId)
        
# delete_meeting_from_excel()
